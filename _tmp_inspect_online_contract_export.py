from openpyxl import load_workbook


ARQUIVO = r"C:\Users\jusgo\Documents\sistema_clinica\_tmp_online_export.xlsx"
CONTRATO_ID = "6753358793211908"


def linhas_sheet(sheet_name: str):
    wb = load_workbook(ARQUIVO, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    header = [str(col or "") for col in rows[0]]
    data = []
    for row in rows[1:]:
        item = {header[idx]: row[idx] for idx in range(len(header))}
        data.append(item)
    return data


def main() -> None:
    contratos = [row for row in linhas_sheet("contratos") if str(row.get("id") or "") == CONTRATO_ID]
    procedimentos = [row for row in linhas_sheet("procedimentos_contrato") if str(row.get("contrato_id") or "") == CONTRATO_ID]
    dentes = [row for row in linhas_sheet("procedimentos_dente") if str(row.get("contrato_id") or "") == CONTRATO_ID]
    print("CONTRATOS", contratos)
    print("PROCEDIMENTOS", procedimentos)
    print("DENTES", dentes)


if __name__ == "__main__":
    main()
